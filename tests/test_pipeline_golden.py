import pandas as pd
from openpyxl import load_workbook
from app.core.engine import process
from app.core.excel_writer import write_excel
from app.core.settings import Preset, Rule, Filter


def test_full_pipeline_end_to_end(tmp_path):
    df = pd.DataFrame({
        "번호": [1, 2, 3, 4],
        "지사": ["강릉", "원주", "강릉", None],
        "공사명": ["활선 작업", "변전 점검", "일반 보수", "취소건"],
        "담당자": ["김", "이", "박", "최"],
    })
    preset = Preset(
        name="강원본부 기본", department_code="4200",
        filters=[Filter(column="지사", op="not_null")],   # None 행 제거 → 3행
        rules=[
            Rule(column="공사명", keyword="활선", match="contains", priority=1, color="#FFFF00"),
            Rule(column="공사명", keyword="변전", match="contains", priority=2, color="#F7B9AF"),
        ],
        drop_columns=["담당자"],
        sheet_split_column="지사",
        sort="priority",
    )

    processed = process(df, preset)
    # 필터: 4→3행, drop: 담당자 제거, 우선순위 라벨 존재
    assert len(processed) == 3
    assert "담당자" not in processed.columns
    assert processed.iloc[0]["점검순위"] == "1순위"

    out = tmp_path / "결과.xlsx"
    write_excel(processed, str(out), sheet_split_column="지사")
    wb = load_workbook(out)
    assert "전체" in wb.sheetnames
    assert "강릉" in wb.sheetnames and "원주" in wb.sheetnames
    ws = wb["전체"]
    headers = [c.value for c in ws[1]]
    assert "_row_color" not in headers and "담당자" not in headers
