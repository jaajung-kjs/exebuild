import pandas as pd
from openpyxl import load_workbook
from app.core.excel_writer import write_excel


def _processed_df():
    return pd.DataFrame({
        "지사": ["강릉", "강릉", "원주"],
        "공사명": ["활선", "일반", "변전"],
        "점검순위": ["1순위", "3순위", "2순위"],
        "_row_color": ["#FFFF00", "", "#F7B9AF"],
    })


def test_writes_file_and_splits_sheets(tmp_path):
    out = tmp_path / "결과.xlsx"
    write_excel(_processed_df(), str(out), sheet_split_column="지사")
    assert out.exists()
    wb = load_workbook(out)
    # 메인 시트 + 지사 고유값 2개(강릉, 원주)
    assert "강릉" in wb.sheetnames
    assert "원주" in wb.sheetnames


def test_internal_color_column_excluded(tmp_path):
    out = tmp_path / "결과.xlsx"
    write_excel(_processed_df(), str(out), sheet_split_column="지사")
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    assert "_row_color" not in headers
    assert "점검순위" in headers


def test_no_split_column_single_sheet(tmp_path):
    out = tmp_path / "결과.xlsx"
    write_excel(_processed_df(), str(out), sheet_split_column="")
    wb = load_workbook(out)
    assert len(wb.sheetnames) == 1


def test_priority_row_fill_applied(tmp_path):
    out = tmp_path / "결과.xlsx"
    write_excel(_processed_df(), str(out), sheet_split_column="")
    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    # 첫 데이터 행(2행)은 1순위 → 노란색 채움
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb.endswith("FFFF00")


def test_split_sheet_fill_aligns_to_filtered_rows(tmp_path):
    # 분할 시트의 배경색이 그 시트에 실제로 들어간 행과 정렬돼야 함.
    # "강릉" 시트: 활선(1순위, 노랑) + 일반(3순위, 무색) 순서.
    # "원주" 시트: 변전(2순위, 분홍) 한 행.
    out = tmp_path / "결과.xlsx"
    write_excel(_processed_df(), str(out), sheet_split_column="지사")
    wb = load_workbook(out)

    gangneung = wb["강릉"]
    # 2행(활선/1순위) → 노랑, 3행(일반/3순위) → 채움 없음
    assert gangneung.cell(row=2, column=1).fill.fgColor.rgb.endswith("FFFF00")
    row3_fill = gangneung.cell(row=3, column=1).fill
    assert not (row3_fill.patternType == "solid" and row3_fill.fgColor.rgb.endswith("FFFF00"))

    wonju = wb["원주"]
    assert wonju.cell(row=2, column=1).fill.fgColor.rgb.endswith("F7B9AF")
