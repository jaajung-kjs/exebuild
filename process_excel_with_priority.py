#!/usr/bin/env python3
"""
Excel RPA 프로그램 - 점검순위 자동 분류
사전신고정보 Excel 파일을 읽어서 분류표.xlsx 기반으로 점검순위 자동 할당
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime, timedelta
import os


def load_classification_keywords(classification_file):
    """
    분류표.xlsx 파일에서 키워드 동적 로드

    Returns:
        tuple: (keywords, special_rules)
        keywords: {
            'A': ['키워드1', '키워드2', ...],  # 2순위(배전)
            'B': ['키워드1', '키워드2', ...],  # 1순위(배전)
            'C': ['키워드1', '키워드2', ...],  # 2순위(송변전)
            'D': ['키워드1', '키워드2', ...],  # 1순위(송변전)
            'E': ['키워드1', '키워드2', ...],  # 2순위(토건,ICT,기타)
            'F': ['키워드1', '키워드2', ...]   # 1순위(토건,ICT,기타)
        }
        special_rules: [
            {'column': 'D', 'keyword': '전력지사', 'priority': '3순위'},
            {'column': 'S', 'keyword': '포함', 'priority': '1순위'},
            ...
        ]
    """
    print(f"📋 분류표 로드 중: {classification_file}")

    try:
        print(f"  [DEBUG] 분류표 파일 존재 확인: {os.path.exists(classification_file)}")
        print(f"  [DEBUG] 분류표 파일 크기: {os.path.getsize(classification_file)} bytes")

        print(f"  [DEBUG] openpyxl.load_workbook() 시작...")
        wb = load_workbook(classification_file)
        print(f"  [DEBUG] 분류표 파일 로드 성공")

        # 첫 번째 시트 명시적으로 선택
        ws = wb.worksheets[0]
        print(f"  시트명: {ws.title}")
        print(f"  [DEBUG] 시트 크기: {ws.max_row} 행 x {ws.max_column} 열")
    except Exception as e:
        print(f"  ❌ [ERROR] 분류표 로드 실패: {type(e).__name__}: {e}")
        raise

    keywords = {}

    # A~F 열의 키워드 수집 (2행부터, 1행은 헤더)
    for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=1):
        col_keywords = []

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value and str(cell_value).strip():
                # 개행 문자가 있는 경우 여러 키워드로 분리
                keywords_in_cell = str(cell_value).strip().split('\n')
                for kw in keywords_in_cell:
                    kw = kw.strip()
                    if kw:
                        col_keywords.append(kw)

        keywords[col_letter] = col_keywords
        header = ws.cell(row=1, column=col_idx).value
        print(f"  {col_letter}열 ({header}): {len(col_keywords)}개 키워드")

    # I, J, K 열에서 특별 우선순위 규칙 수집 (순서대로 적용)
    special_rules = []
    col_i_idx = 9   # I열: 대상 열
    col_j_idx = 10  # J열: 키워드
    col_k_idx = 11  # K열: 순위

    print(f"\n🔥 특별 우선순위 규칙:")
    for row in range(2, ws.max_row + 1):
        target_col = ws.cell(row=row, column=col_i_idx).value
        keyword = ws.cell(row=row, column=col_j_idx).value
        priority = ws.cell(row=row, column=col_k_idx).value

        if target_col and keyword and priority:
            target_col = str(target_col).strip()
            keyword = str(keyword).strip()
            priority_str = f"{priority}순위"

            special_rules.append({
                'column': target_col,
                'keyword': keyword,
                'priority': priority_str
            })
            print(f"  규칙 {len(special_rules)}: {target_col}열에 '{keyword}' 포함 → {priority_str}")

    wb.close()
    return keywords, special_rules


def contains_keyword(text, keyword_list):
    """텍스트에 키워드 리스트 중 하나라도 포함되어 있는지 확인"""
    if not text:
        return False

    text = str(text).strip()
    for keyword in keyword_list:
        if keyword in text:
            return True
    return False


def col_letter_to_index(letter):
    """열 문자를 인덱스로 변환 (A=1, B=2, ..., Z=26, AA=27, ...)"""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def determine_priority(row_data, keywords, special_rules):
    """
    행 데이터를 기반으로 점검순위 결정

    Args:
        row_data: dict with all column data (key: column letter, value: cell value)
        keywords: 분류표에서 로드한 키워드 딕셔너리
        special_rules: 특별 우선순위 규칙 리스트

    Returns:
        str: '1순위', '2순위', '3순위'
    """
    category = row_data.get('H', '')  # 대분류
    col_f = row_data.get('F', '')     # 공사명
    col_o = row_data.get('O', '')     # 기타

    # F열과 O열을 합쳐서 검색
    search_text = f"{col_f} {col_o}"

    # 기본 분류 규칙 적용
    priority = "3순위"  # 기본값

    # 1. 배전
    if category == "배전":
        # 1순위: B열 키워드
        if contains_keyword(search_text, keywords['B']):
            priority = "1순위"
        # 2순위: A열 키워드
        elif contains_keyword(search_text, keywords['A']):
            priority = "2순위"

    # 2. 송전, 변전, 변전(변환)
    elif category in ["송전", "변전", "변전(변환)"]:
        # 1순위: D열 키워드
        if contains_keyword(search_text, keywords['D']):
            priority = "1순위"
        # 2순위: C열 키워드
        elif contains_keyword(search_text, keywords['C']):
            priority = "2순위"

    # 3. 토건, ICT, 기타
    elif category in ["토건", "ICT", "기타"]:
        # 1순위: F열 키워드
        if contains_keyword(search_text, keywords['F']):
            priority = "1순위"
        # 2순위: E열 키워드
        elif contains_keyword(search_text, keywords['E']):
            priority = "2순위"

    # 특별 우선순위 규칙 적용 (순서대로 적용, 나중 규칙이 우선)
    for rule in special_rules:
        target_col = rule['column']
        keyword = rule['keyword']
        rule_priority = rule['priority']

        # 해당 열의 값 가져오기
        cell_value = row_data.get(target_col, '')

        if not cell_value:
            continue

        cell_value_str = str(cell_value)

        # 키워드가 큰따옴표로 감싸져 있으면 정확히 일치하는 경우만 매칭
        if keyword.startswith('"') and keyword.endswith('"'):
            exact_keyword = keyword.strip('"')
            if cell_value_str == exact_keyword:
                priority = rule_priority
        else:
            # 큰따옴표가 없으면 포함 여부로 매칭 (기존 방식)
            if keyword in cell_value_str:
                priority = rule_priority

    return priority


def get_tomorrow_filename():
    """내일 날짜의 파일명 생성"""
    tomorrow = datetime.now() + timedelta(days=1)
    filename = f"사전신고정보_{tomorrow.strftime('%Y-%m-%d')}.xlsx"
    return filename


def process_excel_file(filepath, classification_file):
    """
    Excel 파일 처리
    1. P~T, W~Y, AB~AE 열 삭제
    2. 마지막 열에 '점검순위' 추가
    3. 1행에 필터 추가
    4. 분류표 기반으로 점검순위 자동 할당
    """
    print(f"\n📊 파일 로드 중: {filepath}")

    try:
        print(f"[DEBUG] 입력 파일 존재 확인: {os.path.exists(filepath)}")
        print(f"[DEBUG] 입력 파일 크기: {os.path.getsize(filepath)} bytes")
    except Exception as e:
        print(f"❌ [ERROR] 파일 확인 실패: {type(e).__name__}: {e}")
        raise

    # 분류표 키워드, 특별 규칙 로드
    print(f"\n[DEBUG] === 1단계: 분류표 로드 시작 ===")
    keywords, special_rules = load_classification_keywords(classification_file)
    print(f"[DEBUG] === 1단계: 분류표 로드 완료 ===\n")

    # 우선순위별 고정 색상 정의
    priority_colors = {
        '1순위': 'FFFF00',  # 노란색
        '2순위': 'F7B9AF'   # 분홍색
    }

    # Excel 파일 로드
    print(f"[DEBUG] === 2단계: 입력 파일 로드 시작 ===")
    try:
        print(f"[DEBUG] openpyxl.load_workbook() 시작...")
        wb = load_workbook(filepath)
        print(f"[DEBUG] 입력 파일 로드 성공")
        ws = wb.active
        print(f"[DEBUG] === 2단계: 입력 파일 로드 완료 ===\n")
    except Exception as e:
        print(f"❌ [ERROR] 입력 파일 로드 실패: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise

    print(f"\n현재 시트: {ws.title}")
    print(f"원본 크기: {ws.max_row} 행 x {ws.max_column} 열")

    # 열 문자를 인덱스로 변환하는 함수
    def col_letter_to_index(letter):
        """A=1, B=2, ..., Z=26, AA=27, ..."""
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    # 삭제할 열 범위 정의 (역순으로 삭제해야 인덱스 꼬임 방지)
    delete_ranges = [
        ('AB', 'AE'),  # 28~31
        ('W', 'Y'),    # 23~25
        ('P', 'T'),    # 16~20
    ]

    # 열 삭제
    print("\n[DEBUG] === 3단계: 열 삭제 시작 ===")
    print("🗑️  열 삭제 중...")
    for start_col, end_col in delete_ranges:
        start_idx = col_letter_to_index(start_col)
        end_idx = col_letter_to_index(end_col)
        num_cols = end_idx - start_idx + 1

        print(f"  {start_col}~{end_col} 열 삭제 (인덱스 {start_idx}~{end_idx}, {num_cols}개 열)")
        ws.delete_cols(start_idx, num_cols)

    print(f"삭제 후 크기: {ws.max_row} 행 x {ws.max_column} 열")

    # 마지막 열에 '점검순위' 추가
    priority_col = ws.max_column + 1
    print(f"\n📌 마지막 열({priority_col})에 '점검순위' 추가")
    ws.cell(row=1, column=priority_col, value="점검순위")

    # 1행에 필터 추가
    print("🔍 1행에 자동 필터 추加")
    ws.auto_filter.ref = ws.dimensions

    # 틀 고정 (1행 고정)
    print("📌 제목행 틀 고정 (Freeze Panes)")
    ws.freeze_panes = 'A2'  # A2 셀 기준으로 고정 (1행 고정)

    # 점검순위 자동 할당 (2행부터)
    print(f"\n⚡ 점검순위 자동 할당 중... (총 {ws.max_row - 1}개 행)")

    # 열 문자를 인덱스로 매핑
    def index_to_col_letter(idx):
        """인덱스를 열 문자로 변환 (1=A, 2=B, ...)"""
        result = ""
        while idx > 0:
            idx -= 1
            result = chr(ord('A') + idx % 26) + result
            idx //= 26
        return result

    priority_counts = {"1순위": 0, "2순위": 0, "3순위": 0}

    for row in range(2, ws.max_row + 1):
        # 현재 행의 모든 열 데이터를 딕셔너리로 수집
        row_data = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = index_to_col_letter(col_idx)
            row_data[col_letter] = ws.cell(row=row, column=col_idx).value

        priority = determine_priority(row_data, keywords, special_rules)
        ws.cell(row=row, column=priority_col, value=priority)
        priority_counts[priority] += 1

        # 진행상황 출력 (100개마다)
        if row % 100 == 0:
            print(f"  진행중... {row}/{ws.max_row} 행")

    print(f"\n📊 점검순위 할당 완료:")
    print(f"  1순위: {priority_counts['1순위']}건")
    print(f"  2순위: {priority_counts['2순위']}건")
    print(f"  3순위: {priority_counts['3순위']}건")

    # A열 기준 오름차순 정렬 (2행부터, 1행은 헤더)
    print(f"\n📊 A열 기준 오름차순 정렬 중...")
    # 데이터 범위 정의
    data_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    ws.auto_filter.ref = data_range
    # 정렬은 수동으로 데이터를 읽어서 재배치
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data_rows.append(row_data)

    # A열(인덱스 0) 기준 오름차순 정렬
    data_rows.sort(key=lambda x: x[0] if x[0] is not None else 0, reverse=False)

    # 정렬된 데이터를 다시 쓰기
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    print("✅ 정렬 완료")

    # 열 너비 적용 (코드 내장) - 스타일 적용 전에 먼저 실행
    print(f"\n📏 열 너비 적용 중...")

    # 고정 열 너비 정의
    column_widths = {
        'A': 6.6640625,
        'B': 9.83203125,
        'C': 8.83203125,
        'D': 13.83203125,
        'E': 12.1640625,
        'F': 21.33203125,
        'G': 14.5,
        'H': 13.0,
        'I': 13.0,
        'J': 17.5,
        'K': 8.5,
        'L': 13.0,
        'M': 13.0,
        'N': 21.83203125,
        'O': 28.33203125,
        'P': 17.1640625,
        'Q': 10.0,
        'R': 10.1640625,
        'S': 10.0,
        'T': 11.5
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        print(f"  {col_letter}열: {width}")

    print("✅ 열 너비 적용 완료")

    # 전체 셀 가운데 정렬 + 텍스트 자동 줄바꿈
    print(f"\n🎨 전체 셀 가운데 정렬 및 자동 줄바꿈 적용 중...")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment

    print("✅ 가운데 정렬 및 자동 줄바꿈 완료")

    # 1행 제목행 스타일 적용 (BOLD + 배경색)
    print(f"\n🎨 제목행 스타일 적용 중...")
    bold_font = Font(bold=True)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = gray_fill

    print("✅ 제목행 스타일 완료")

    # 우선순위별 행 배경색 적용
    print(f"\n🎨 우선순위별 행 배경색 적용 중...")

    priority_col_idx = ws.max_column  # 점검순위 열 인덱스
    color_counts = {"1순위": 0, "2순위": 0}

    for row in range(2, ws.max_row + 1):
        priority = ws.cell(row=row, column=priority_col_idx).value

        if priority in priority_colors:
            color = priority_colors[priority]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # 해당 행의 모든 셀에 배경색 적용
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

            color_counts[priority] += 1

    print(f"  1순위 행: {color_counts['1순위']}개 (색상: {priority_colors.get('1순위', 'N/A')})")
    print(f"  2순위 행: {color_counts['2순위']}개 (색상: {priority_colors.get('2순위', 'N/A')})")
    print("✅ 행 배경색 적용 완료")

    # 모든 셀에 테두리 적용
    print(f"\n🎨 모든 셀에 테두리 적용 중...")
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    print("✅ 테두리 적용 완료")

    # 파일 저장 (새로운 파일명 형식: YYMMDD 공사현장 점검 우선순위 리스트.xlsx)
    print(f"\n[DEBUG] === 최종 단계: 파일 저장 시작 ===")
    tomorrow = datetime.now() + timedelta(days=1)
    output_filename = tomorrow.strftime('%y%m%d') + ' 공사현장 점검 우선순위 리스트.xlsx'
    output_filepath = os.path.join(os.path.dirname(filepath), output_filename)

    print(f"💾 처리된 파일 저장: {output_filename}")
    print(f"[DEBUG] 저장 경로: {output_filepath}")

    try:
        print(f"[DEBUG] wb.save() 시작...")
        wb.save(output_filepath)
        print(f"[DEBUG] 파일 저장 성공")
        print(f"[DEBUG] 저장된 파일 크기: {os.path.getsize(output_filepath)} bytes")
        print(f"[DEBUG] === 최종 단계: 파일 저장 완료 ===\n")
        print("✅ 완료!")
    except Exception as e:
        print(f"❌ [ERROR] 파일 저장 실패: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise

    return output_filepath


def main():
    """메인 함수"""
    # 분류표 파일
    classification_file = "분류표.xlsx"
    if not os.path.exists(classification_file):
        print(f"❌ 분류표 파일을 찾을 수 없습니다: {classification_file}")
        return

    # 내일 날짜 파일명 생성
    filename = get_tomorrow_filename()
    filepath = os.path.join(os.getcwd(), filename)

    # 파일 존재 확인
    if not os.path.exists(filepath):
        print(f"❌ 파일을 찾을 수 없습니다: {filename}")
        print(f"   경로: {filepath}")
        return

    # Excel 파일 처리
    try:
        output_file = process_excel_file(filepath, classification_file)
        print(f"\n🎉 처리 완료: {output_file}")
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
