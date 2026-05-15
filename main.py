"""
KEPCO RPA Main Program
Automated workflow: Download → Process → Send Email

실행 방법:
    python main.py

요구사항:
    - PowerGate 실행 중
    - 분류표.xlsx 파일 존재 (같은 디렉토리)
    - 인터넷 연결 (사내망)
"""

from datetime import datetime, timedelta
import sys
import os
import time

# Import modules
import auth
import downloader
import processor
import mailer
from config import get_classification_file_path, get_output_dir


def read_department_code():
    """
    Read DEPARTMENT_CODE from 분류표.xlsx sheet 1, cell M2

    Returns:
        str: Department code (default: "4200" if read fails)
    """
    try:
        from openpyxl import load_workbook

        classification_file = get_classification_file_path()

        if not os.path.exists(classification_file):
            print(f"⚠️  분류표.xlsx 파일을 찾을 수 없습니다. 기본값 사용: 4200")
            return "4200"

        wb = load_workbook(classification_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # Sheet 1
        dept_code = ws['M2'].value
        wb.close()

        if dept_code is None or str(dept_code).strip() == '':
            print(f"⚠️  M2 셀이 비어있습니다. 기본값 사용: 4200")
            return "4200"

        dept_code_str = str(dept_code).strip()
        print(f"✅ 부서코드 로드: {dept_code_str} (분류표.xlsx M2)")
        return dept_code_str

    except Exception as e:
        print(f"⚠️  부서코드 읽기 실패: {e}")
        print(f"   기본값 사용: 4200")
        return "4200"


def read_target_date():
    """
    Read target date from 분류표.xlsx sheet 1, cell N2

    Returns:
        datetime.date: Target date (default: tomorrow if cell is empty or read fails)
    """
    tomorrow = (datetime.now() + timedelta(days=1)).date()

    try:
        from openpyxl import load_workbook

        classification_file = get_classification_file_path()

        if not os.path.exists(classification_file):
            print(f"⚠️  분류표.xlsx 파일을 찾을 수 없습니다. 기본값 사용: 내일({tomorrow})")
            return tomorrow

        wb = load_workbook(classification_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # Sheet 1
        cell_value = ws['N2'].value
        wb.close()

        if cell_value is None or str(cell_value).strip() == '':
            print(f"✅ 대상 날짜: {tomorrow} (N2 셀 비어있음 → 내일)")
            return tomorrow

        # openpyxl이 datetime 객체를 반환하는 경우
        if hasattr(cell_value, 'date'):
            target = cell_value.date()
            print(f"✅ 대상 날짜 로드: {target} (분류표.xlsx N2)")
            return target

        # 문자열인 경우 YYYY-MM-DD 파싱
        target = datetime.strptime(str(cell_value).strip(), '%Y-%m-%d').date()
        print(f"✅ 대상 날짜 로드: {target} (분류표.xlsx N2)")
        return target

    except Exception as e:
        print(f"⚠️  대상 날짜 읽기 실패: {e}")
        print(f"   기본값 사용: 내일({tomorrow})")
        return tomorrow


def print_header():
    """Print program header"""
    print("\n" + "=" * 60)
    print("KEPCO RPA 자동화 시스템")
    print("=" * 60)
    print(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"출력 경로: {get_output_dir()}")
    print("=" * 60 + "\n")


def print_footer(success):
    """Print program footer"""
    print("\n" + "=" * 60)
    print("작업 완료")
    print("=" * 60)
    print(f"종료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"결과: {'✅ 성공' if success else '❌ 실패'}")
    print("=" * 60 + "\n")


def main():
    """Main workflow execution"""
    success = False

    try:
        print_header()

        # Read target date and department code
        target_date = read_target_date()
        dept_code = read_department_code()

        # Prepare date formats for pipeline
        target_date_str = target_date.strftime('%Y-%m-%d')         # downloader용
        target_date_yymmdd = target_date.strftime('%y%m%d')        # processor, mailer subject
        target_date_yy_mm_dd = target_date.strftime("'%y-%m-%d")   # mailer body

        print(f"\n📅 대상 날짜: {target_date_str} (YYMMDD: {target_date_yymmdd})")

        # ============================================
        # Step 1: Authentication + Download (with retry)
        # ============================================
        MAX_RETRIES = 10
        RETRY_DELAY = 3
        print(f"\n⚡ [1/3] 인증 + 다운로드 (최대 {MAX_RETRIES}회 시도)")
        print("-" * 60)

        session = None
        df = None

        for attempt in range(1, MAX_RETRIES + 1):
            if attempt > 1:
                print(f"\n🔄 재시도 {attempt}/{MAX_RETRIES} ({RETRY_DELAY}초 대기 후)...")
                time.sleep(RETRY_DELAY)

            # 1. Authentication (auth 내부에서 GET 검증 + 재시도)
            session = auth.authenticate()
            if session is None:
                print(f"  ⚠️  인증 실패")
                continue

            # 2. Download
            df = downloader.download_excel_to_dataframe(
                session, date_from=target_date_str, department_code=dept_code
            )
            if df is not None:
                break

            print(f"  ⚠️  다운로드 실패 — 재인증 후 재시도합니다")
        else:
            print(f"\n❌ {MAX_RETRIES}회 시도 후에도 인증+다운로드에 실패했습니다")
            print("\n해결 방법:")
            print("  1. PowerGate가 실행 중인지 확인하세요")
            print("  2. 사내망에 연결되어 있는지 확인하세요")
            print("  3. Work Monitor 서버 접근 권한을 확인하세요")
            print("  4. 해당 날짜에 데이터가 있는지 확인하세요")
            return

        print(f"\n  ✅ 인증 + 다운로드 성공")

        # ============================================
        # Step 2: Load classification and process data
        # ============================================
        print(f"\n⚡ [2/3] 데이터 가공")
        print("-" * 60)

        classification_file = get_classification_file_path()

        if not os.path.exists(classification_file):
            print(f"\n❌ 분류표 파일을 찾을 수 없습니다")
            print(f"   경로: {classification_file}")
            print("\n해결 방법:")
            print("  1. 분류표.xlsx 파일이 프로그램과 같은 폴더에 있는지 확인하세요")
            return

        # Load classification keywords and mail config
        keywords, special_rules, mail_config = processor.load_classification_and_mail_config(
            classification_file
        )

        if mail_config is None:
            print("\n❌ 메일 설정을 찾을 수 없습니다")
            print("\n해결 방법:")
            print("  1. 분류표.xlsx 파일에 시트 2가 있는지 확인하세요")
            print("  2. 시트 2에 메일 설정 정보가 올바르게 입력되어 있는지 확인하세요")
            return

        # 서울본부 위험순위 룰 + 협력사 제재현황 로드 (시트 없으면 비활성)
        seoul_rules = processor.load_seoul_rules(classification_file)
        sanctions = processor.load_sanctions(classification_file)

        # Process DataFrame and save to Excel
        output_file = processor.process_dataframe(
            df, keywords, special_rules,
            target_date_yymmdd=target_date_yymmdd,
            seoul_rules=seoul_rules,
            sanctions=sanctions,
        )

        if output_file is None or not os.path.exists(output_file):
            print("\n❌ 데이터 가공 실패")
            return

        print(f"\n✅ 가공 완료: {os.path.basename(output_file)}")

        # ============================================
        # Step 4: Send email
        # ============================================
        print(f"\n⚡ [3/3] 메일 전송")
        print("-" * 60)

        result = mailer.send_bizmail(
            session=session,
            mail_config=mail_config,
            attachment_paths=[output_file],
            date_yymmdd=target_date_yymmdd,
            date_yy_mm_dd=target_date_yy_mm_dd
        )

        if result['success']:
            print("\n✅ 모든 작업이 성공적으로 완료되었습니다!")
            print(f"\n최종 결과물:")
            print(f"  파일: {os.path.basename(output_file)}")
            print(f"  위치: {output_file}")
            print(f"\n메일 발송:")
            print(f"  수신자: {len(mail_config['recipients'])}명")
            print(f"  제목: {mail_config['subject']}")
            success = True
        else:
            print(f"\n❌ 메일 전송 실패: {result['message']}")
            print(f"\n파일은 저장되었습니다:")
            print(f"  {output_file}")

    except KeyboardInterrupt:
        print("\n\n⚠️  사용자가 작업을 중단했습니다")

    except Exception as e:
        print(f"\n❌ 예상치 못한 오류가 발생했습니다:")
        print(f"   {str(e)}")
        print("\n상세 오류:")
        import traceback
        traceback.print_exc()

    finally:
        print_footer(success)


if __name__ == '__main__':
    """
    KEPCO RPA 자동화 프로그램

    기능:
        1. Work Monitor에서 사전신고정보 Excel 다운로드
        2. 분류표 기준으로 점검순위 자동 분류
        3. BizMail로 자동 발송

    설정:
        - config.py: 기본 설정 (날짜, 부서코드 등)
        - 분류표.xlsx:
            * 시트 1: 분류 키워드 규칙
            * 시트 2: 메일 설정 (발신자, 수신자, 제목, 본문)

    실행 환경:
        - PowerGate 실행 필요
        - KEPCO 사내망 접속 필요
        - 분류표.xlsx 파일 필요 (같은 디렉토리)

    출력:
        - 파일 위치: 프로그램과 같은 폴더
        - 파일명: YYMMDD 공사현장 점검 우선순위 리스트.xlsx
    """
    try:
        main()
    except Exception as e:
        print(f"\n프로그램 실행 중 치명적 오류 발생: {e}")
        sys.exit(1)
