"""엑셀 생성기 — 가공된 DataFrame → 지사별 다중 시트 + 서식 xlsx"""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INTERNAL_COLOR_COL = "_row_color"

MIN_COL_W = 8      # 최소 열 너비
MAX_COL_W = 55     # 최대 열 너비(이 이상은 안 늘리고 자동 줄바꿈)


def _disp_len(v) -> int:
    """표시 폭 추정 — 한글/CJK는 2칸, 그 외는 1칸."""
    return sum(2 if ord(ch) >= 0x1100 else 1 for ch in str(v))


def _safe_sheet_name(value: str) -> str:
    name = str(value)[:31]
    for ch in '/\\*[]:?':
        name = name.replace(ch, "_")
    return name or "시트"


def _hex(color: str) -> str:
    """'#FFFF00' → 'FFFF00' (openpyxl용). 빈 값이면 ''."""
    return color.lstrip("#").upper() if color else ""


def _apply_format(ws, colors: list[str], n_cols: int):
    """한 시트에 서식 적용. colors[i]는 데이터 i번째 행(엑셀 i+2행)의 배경색 헥사."""
    # 전체 가운데 정렬 + 자동 줄바꿈(열 최대너비 초과 시 줄바꿈, 행 높이는 Excel 자동)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=n_cols):
        for cell in row:
            cell.alignment = center_wrap

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = header_fill

    for i, hexcolor in enumerate(colors):
        if not hexcolor:
            continue
        fill = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")
        for col in range(1, n_cols + 1):
            ws.cell(row=i + 2, column=col).fill = fill

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = border

    # 열 너비 자동 조절 — 내용 길이(한글 폭 반영) 기준, 최소~최대 클램프
    for col in range(1, n_cols + 1):
        maxlen = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                dl = _disp_len(v)
                if dl > maxlen:
                    maxlen = dl
        width = min(MAX_COL_W, max(MIN_COL_W, maxlen * 1.15 + 2))
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_excel(df: pd.DataFrame, output_path: str, sheet_split_column: str = "") -> str:
    colors_all = (
        df[INTERNAL_COLOR_COL].fillna("").tolist()
        if INTERNAL_COLOR_COL in df.columns else [""] * len(df)
    )
    visible = df.drop(columns=[INTERNAL_COLOR_COL], errors="ignore")
    n_cols = len(visible.columns)

    split = sheet_split_column if sheet_split_column in visible.columns else ""

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 메인 시트 (전체)
        visible.to_excel(writer, index=False, sheet_name="전체")

        sheet_colors = {"전체": [_hex(c) for c in colors_all]}

        if split:
            for value in sorted(visible[split].dropna().unique()):
                mask = (visible[split] == value).tolist()
                part = visible[visible[split] == value]
                part_colors = [_hex(c) for c, keep in zip(colors_all, mask) if keep]
                name = _safe_sheet_name(value)
                part.to_excel(writer, index=False, sheet_name=name)
                sheet_colors[name] = part_colors

        wb = writer.book
        for name, colors in sheet_colors.items():
            _apply_format(wb[name], colors, n_cols)

    return output_path
