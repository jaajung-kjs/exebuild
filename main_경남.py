"""
KEPCO RPA Main Program — 경남본부 전용 변형판

뼈대(기존 RPA)와 차이점:
    - 서울본부 위험순위 룰 / 제재현황 로드 안 함 (Seoul 로직 비활성)
    - 컬럼 drop 없음 (다운로드된 31개 컬럼 전부 보존)

그 외 동작은 main.py 와 동일.
"""

from datetime import datetime, timedelta
import sys
import os
import time

import auth
import downloader
import processor
import mailer
from config import get_classification_file_path, get_output_dir


def read_department_code():
    try:
        from openpyxl import load_workbook
        classification_file = get_classification_file_path()
        if not os.path.exists(classification_file):
            print(f"⚠️  분류표.xlsx 파일을 찾을 수 없습니다. 기본값 사용: 4200")
            return "4200"
        wb = load_workbook(classification_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        dept_code = ws['M2'].value
        wb.close()
        if dept_code is None or str(dept_code).strip() == '':
            print(f"⚠️  M2 셀이 비어있습니다. 기본값 사용: 4200")
            return "4200"
        dept_code_str = str(dept_code).strip()
        print(f"✅ 부서코드 로드: {dept_code_str} (분류표.xlsx M2)")
        return dept_code_str
    except Exception as e:
        print(f"⚠️  부서코드 읽기 실패: {e}")
        print(f"   기본값 사용: 4200")
        return "4200"


def read_target_date():
    tomorrow = (datetime.now() + timedelta(days=1)).date()
    try:
        from openpyxl import load_workbook
        classification_file = get_classification_file_path()
        if not os.path.exists(classification_file):
            print(f"⚠️  분류표.xlsx 파일을 찾을 수 없습니다. 기본값 사용: 내일({tomorrow})")
            return tomorrow
        wb = load_workbook(classification_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        cell_value = ws['N2'].value
        wb.close()
        if cell_value is None or str(cell_value).strip() == '':
            print(f"✅ 대상 날짜: {tomorrow} (N2 셀 비어있음 → 내일)")
            return tomorrow
        if hasattr(cell_value, 'date'):
            target = cell_value.date()
            print(f"✅ 대상 날짜 로드: {target} (분류표.xlsx N2)")
            return target
        target = datetime.strptime(str(cell_value).strip(), '%Y-%m-%d').date()
        print(f"✅ 대상 날짜 로드: {target} (분류표.xlsx N2)")
        return target
    except Exception as e:
        print(f"⚠️  대상 날짜 읽기 실패: {e}")
        print(f"   기본값 사용: 내일({tomorrow})")
        return tomorrow


def print_header():
    print("\n" + "=" * 60)
    print("KEPCO RPA 자동화 시스템 — 경남본부 전용")
    print("=" * 60)
    print(f"시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"출력 경로: {get_output_dir()}")
    print("=" * 60 + "\n")


def print_footer(success):
    print("\n" + "=" * 60)
    print("작업 완료")
    print("=" * 60)
    print(f"종료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"결과: {'✅ 성공' if success else '❌ 실패'}")
    print("=" * 60 + "\n")


def main():
    success = False
    try:
        print_header()

        target_date = read_target_date()
        dept_code = read_department_code()

        target_date_str = target_date.strftime('%Y-%m-%d')
        target_date_yymmdd = target_date.strftime('%y%m%d')
        target_date_yy_mm_dd = target_date.strftime("'%y-%m-%d")

        print(f"\n📅 대상 날짜: {target_date_str} (YYMMDD: {target_date_yymmdd})")

        MAX_RETRIES = 10
        RETRY_DELAY = 3
        print(f"\n⚡ [1/3] 인증 + 다운로드 (최대 {MAX_RETRIES}회 시도)")
        print("-" * 60)

        session = None
        df = None

        for attempt in range(1, MAX_RETRIES + 1):
            if attempt > 1:
                print(f"\n🔄 재시도 {attempt}/{MAX_RETRIES} ({RETRY_DELAY}초 대기 후)...")
                time.sleep(RETRY_DELAY)

            session = auth.authenticate()
            if session is None:
                print(f"  ⚠️  인증 실패")
                continue

            df = downloader.download_excel_to_dataframe(
                session, date_from=target_date_str, department_code=dept_code
            )
            if df is not None:
                break

            print(f"  ⚠️  다운로드 실패 — 재인증 후 재시도합니다")
        else:
            print(f"\n❌ {MAX_RETRIES}회 시도 후에도 인증+다운로드에 실패했습니다")
            return

        print(f"\n  ✅ 인증 + 다운로드 성공")

        print(f"\n⚡ [2/3] 데이터 가공")
        print("-" * 60)

        classification_file = get_classification_file_path()
        if not os.path.exists(classification_file):
            print(f"\n❌ 분류표 파일을 찾을 수 없습니다: {classification_file}")
            return

        keywords, special_rules, mail_config = processor.load_classification_and_mail_config(
            classification_file
        )
        if mail_config is None:
            print("\n❌ 메일 설정을 찾을 수 없습니다")
            return

        # 경남본부 변형:
        #   - 서울본부 룰/제재현황 로드 안 함
        #   - drop_columns_override=[] → 전 컬럼 보존
        output_file = processor.process_dataframe(
            df, keywords, special_rules,
            target_date_yymmdd=target_date_yymmdd,
            drop_columns_override=[],
        )

        if output_file is None or not os.path.exists(output_file):
            print("\n❌ 데이터 가공 실패")
            return

        print(f"\n✅ 가공 완료: {os.path.basename(output_file)}")

        print(f"\n⚡ [3/3] 메일 전송")
        print("-" * 60)

        result = mailer.send_bizmail(
            session=session,
            mail_config=mail_config,
            attachment_paths=[output_file],
            date_yymmdd=target_date_yymmdd,
            date_yy_mm_dd=target_date_yy_mm_dd
        )

        if result['success']:
            print("\n✅ 모든 작업이 성공적으로 완료되었습니다!")
            success = True
        else:
            print(f"\n❌ 메일 전송 실패: {result['message']}")
            print(f"파일은 저장되었습니다: {output_file}")

    except KeyboardInterrupt:
        print("\n\n⚠️  사용자가 작업을 중단했습니다")
    except Exception as e:
        print(f"\n❌ 예상치 못한 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print_footer(success)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n프로그램 실행 중 치명적 오류 발생: {e}")
        sys.exit(1)
