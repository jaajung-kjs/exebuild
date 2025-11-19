"""
Excel processing module with priority classification
Reads classification rules and mail config from 분류표.xlsx
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime, timedelta
import os
import pandas as pd
from config import get_classification_file_path, get_output_dir


def load_classification_and_mail_config(classification_file):
    """
    Load classification keywords from sheet 1 and mail config from sheet 2

    Returns:
        tuple: (keywords, special_rules, mail_config)

        keywords: {
            'A': ['키워드1', '키워드2', ...],  # 2순위(배전)
            'B': ['키워드1', '키워드2', ...],  # 1순위(배전)
            'C': ['키워드1', '키워드2', ...],  # 2순위(송변전)
            'D': ['키워드1', '키워드2', ...],  # 1순위(송변전)
            'E': ['키워드1', '키워드2', ...],  # 2순위(토건,ICT,기타)
            'F': ['키워드1', '키워드2', ...]   # 1순위(토건,ICT,기타)
        }

        special_rules: [
            {'column': 'D', 'keyword': '전력지사', 'priority': '3순위'},
            ...
        ]

        mail_config: {
            'from_name': str,       # A2: 작성자 이름
            'from_email': str,      # B2: 작성자 이메일
            'recipients': [str],    # D2~DN: 수신자 이메일 리스트
            'subject': str,         # E2: 메일 제목
            'body': str            # F2: 메일 본문
        }
    """
    print("=" * 60)
    print("분류표 로드")
    print("=" * 60)

    try:
        wb = load_workbook(classification_file)
        print(f"  파일 로드 성공: {classification_file}")
    except Exception as e:
        print(f"  [오류] 파일 로드 실패: {e}")
        raise

    # ========================================
    # Sheet 1: Classification keywords
    # ========================================
    ws1 = wb.worksheets[0]
    print(f"\n[시트 1] {ws1.title}")
    print(f"  크기: {ws1.max_row} 행 x {ws1.max_column} 열")

    keywords = {}

    # A~F 열의 키워드 수집 (2행부터, 1행은 헤더)
    for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], start=1):
        col_keywords = []

        for row in range(2, ws1.max_row + 1):
            cell_value = ws1.cell(row=row, column=col_idx).value
            if cell_value and str(cell_value).strip():
                # 개행 문자가 있는 경우 여러 키워드로 분리
                keywords_in_cell = str(cell_value).strip().split('\n')
                for kw in keywords_in_cell:
                    kw = kw.strip()
                    if kw:
                        col_keywords.append(kw)

        keywords[col_letter] = col_keywords
        header = ws1.cell(row=1, column=col_idx).value
        print(f"  {col_letter}열 ({header}): {len(col_keywords)}개 키워드")

    # I, J, K 열에서 특별 우선순위 규칙 수집
    special_rules = []
    col_i_idx = 9   # I열: 대상 열
    col_j_idx = 10  # J열: 키워드
    col_k_idx = 11  # K열: 순위

    print(f"\n  특별 우선순위 규칙:")
    for row in range(2, ws1.max_row + 1):
        target_col = ws1.cell(row=row, column=col_i_idx).value
        keyword = ws1.cell(row=row, column=col_j_idx).value
        priority = ws1.cell(row=row, column=col_k_idx).value

        if target_col and keyword and priority:
            target_col = str(target_col).strip()
            keyword = str(keyword).strip()
            priority_str = f"{priority}순위"

            special_rules.append({
                'column': target_col,
                'keyword': keyword,
                'priority': priority_str
            })
            print(f"    규칙 {len(special_rules)}: {target_col}열에 '{keyword}' 포함 → {priority_str}")

    # ========================================
    # Sheet 2: Mail configuration
    # ========================================
    ws2 = wb.worksheets[1] if len(wb.worksheets) >= 2 else None

    mail_config = None

    if ws2:
        print(f"\n[시트 2] {ws2.title}")
        print(f"  크기: {ws2.max_row} 행 x {ws2.max_column} 열")

        # A2: 작성자 이름
        from_name = ws2.cell(row=2, column=1).value

        # B2: 작성자 이메일
        from_email = ws2.cell(row=2, column=2).value

        # D2~DN: 수신자 이메일 (빈 셀 나올 때까지)
        recipients = []
        row_num = 2
        while True:
            email = ws2.cell(row=row_num, column=4).value
            if email is None or str(email).strip() == '':
                break
            recipients.append(str(email).strip())
            row_num += 1

        # E2: 메일 제목
        subject = ws2.cell(row=2, column=5).value

        # F2: 메일 본문 (HTML)
        body = ws2.cell(row=2, column=6).value

        mail_config = {
            'from_name': str(from_name).strip() if from_name else '',
            'from_email': str(from_email).strip() if from_email else '',
            'recipients': recipients,
            'subject': str(subject).strip() if subject else '',
            'body': str(body).strip() if body else ''
        }

        print(f"\n  메일 설정:")
        print(f"    발신자: {mail_config['from_name']} <{mail_config['from_email']}>")
        print(f"    수신자: {len(mail_config['recipients'])}명")
        for idx, rcpt in enumerate(mail_config['recipients'], 1):
            print(f"      {idx}. {rcpt}")
        print(f"    제목: {mail_config['subject']}")
        print(f"    본문: {mail_config['body'][:50]}..." if len(mail_config['body']) > 50 else f"    본문: {mail_config['body']}")
    else:
        print(f"\n  [경고] 시트 2가 없습니다. 기본 메일 설정 사용")

    wb.close()
    print("=" * 60 + "\n")

    return keywords, special_rules, mail_config


def contains_keyword(text, keyword_list):
    """텍스트에 키워드 리스트 중 하나라도 포함되어 있는지 확인"""
    if not text:
        return False

    text = str(text).strip()
    for keyword in keyword_list:
        if keyword in text:
            return True
    return False


def determine_priority(row_data, keywords, special_rules):
    """
    행 데이터를 기반으로 점검순위 결정

    Args:
        row_data: dict with all column data (key: column letter, value: cell value)
        keywords: 분류표에서 로드한 키워드 딕셔너리
        special_rules: 특별 우선순위 규칙 리스트

    Returns:
        str: '1순위', '2순위', '3순위'
    """
    category = row_data.get('H', '')  # 대분류
    col_f = row_data.get('F', '')     # 공사명
    col_o = row_data.get('O', '')     # 기타

    # F열과 O열을 합쳐서 검색
    search_text = f"{col_f} {col_o}"

    # 기본 분류 규칙 적용
    priority = "3순위"  # 기본값

    # 1. 배전
    if category == "배전":
        if contains_keyword(search_text, keywords['B']):
            priority = "1순위"
        elif contains_keyword(search_text, keywords['A']):
            priority = "2순위"

    # 2. 송전, 변전, 변전(변환)
    elif category in ["송전", "변전", "변전(변환)"]:
        if contains_keyword(search_text, keywords['D']):
            priority = "1순위"
        elif contains_keyword(search_text, keywords['C']):
            priority = "2순위"

    # 3. 토건, ICT, 기타
    elif category in ["토건", "ICT", "기타"]:
        if contains_keyword(search_text, keywords['F']):
            priority = "1순위"
        elif contains_keyword(search_text, keywords['E']):
            priority = "2순위"

    # 특별 우선순위 규칙 적용 (순서대로 적용, 나중 규칙이 우선)
    for rule in special_rules:
        target_col = rule['column']
        keyword = rule['keyword']
        rule_priority = rule['priority']

        # 해당 열의 값 가져오기
        cell_value = row_data.get(target_col, '')

        if not cell_value:
            continue

        cell_value_str = str(cell_value)

        # 키워드가 큰따옴표로 감싸져 있으면 정확히 일치하는 경우만 매칭
        if keyword.startswith('"') and keyword.endswith('"'):
            exact_keyword = keyword.strip('"')
            if cell_value_str == exact_keyword:
                priority = rule_priority
        else:
            # 큰따옴표가 없으면 포함 여부로 매칭
            if keyword in cell_value_str:
                priority = rule_priority

    return priority


def index_to_col_letter(idx):
    """인덱스를 열 문자로 변환 (1=A, 2=B, ...)"""
    result = ""
    while idx > 0:
        idx -= 1
        result = chr(ord('A') + idx % 26) + result
        idx //= 26
    return result


def col_letter_to_index(letter):
    """열 문자를 인덱스로 변환 (A=1, B=2, ...)"""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def process_dataframe(df, keywords, special_rules):
    """
    Process DataFrame and save as Excel with priority classification

    Args:
        df: pandas DataFrame from download
        keywords: classification keywords
        special_rules: special priority rules

    Returns:
        str: Output file path
    """
    print("=" * 60)
    print("데이터 가공")
    print("=" * 60)

    print(f"\n  원본 크기: {len(df)} 행 x {len(df.columns)} 열")

    # 1. 불필요한 열 삭제 (DataFrame에서 직접)
    print(f"\n  불필요한 열 삭제 중...")

    # 열 인덱스로 삭제할 범위 계산 (0-based index)
    # P~T: 15~19, W~Y: 22~24, AB~AE: 27~30
    cols_to_drop = []

    # P(16열) ~ T(20열)
    if len(df.columns) >= 20:
        cols_to_drop.extend(df.columns[15:20].tolist())

    # W(23열) ~ Y(25열)
    if len(df.columns) >= 25:
        cols_to_drop.extend(df.columns[22:25].tolist())

    # AB(28열) ~ AE(31열)
    if len(df.columns) >= 31:
        cols_to_drop.extend(df.columns[27:31].tolist())

    df = df.drop(columns=cols_to_drop, errors='ignore')
    print(f"  삭제 후 크기: {len(df)} 행 x {len(df.columns)} 열")

    # 2. 점검순위 열 추가 및 자동 할당
    print(f"\n  점검순위 자동 할당 중... (총 {len(df)}개 행)")

    priorities = []
    priority_counts = {"1순위": 0, "2순위": 0, "3순위": 0}

    for idx, row in df.iterrows():
        # 행 데이터를 열 문자(A, B, C, ...) 기준 딕셔너리로 변환
        row_data = {}
        for col_idx, col_name in enumerate(df.columns):
            col_letter = index_to_col_letter(col_idx + 1)
            row_data[col_letter] = row[col_name]

        priority = determine_priority(row_data, keywords, special_rules)
        priorities.append(priority)
        priority_counts[priority] += 1

        if (idx + 1) % 100 == 0:
            print(f"    진행중... {idx + 1}/{len(df)} 행")

    # 점검순위 열 추가
    df['점검순위'] = priorities

    print(f"\n  점검순위 할당 완료:")
    print(f"    1순위: {priority_counts['1순위']}건")
    print(f"    2순위: {priority_counts['2순위']}건")
    print(f"    3순위: {priority_counts['3순위']}건")

    # 3. A열 기준 오름차순 정렬
    print(f"\n  첫 번째 열 기준 정렬 중...")
    df = df.sort_values(by=df.columns[0], ascending=True)

    # 4. 최종 파일 저장 (openpyxl로 서식 적용)
    print(f"\n  Excel 서식 적용 및 저장 중...")

    output_dir = get_output_dir()
    tomorrow = datetime.now() + timedelta(days=1)
    output_filename = tomorrow.strftime('%y%m%d') + ' 공사현장 점검 우선순위 리스트.xlsx'
    output_filepath = os.path.join(output_dir, output_filename)

    # D열(4번째 열)의 고유값 추출
    d_column_name = df.columns[3]  # 0-based index, D열은 4번째
    unique_d_values = sorted(df[d_column_name].dropna().unique())

    print(f"\n  D열({d_column_name}) 고유값: {len(unique_d_values)}개")
    for idx, val in enumerate(unique_d_values, 1):
        count = len(df[df[d_column_name] == val])
        print(f"    {idx}. {val}: {count}건")

    # DataFrame을 Excel로 저장 (다중 시트)
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        # 메인 시트명: C2 셀 값 사용 (C열은 3번째 열, 인덱스 2)
        main_sheet_name = str(df.iloc[0, 2]) if len(df) > 0 else '메인'

        # 1. 메인 시트 생성 (전체 데이터)
        print(f"\n  메인 시트 생성 중: {main_sheet_name}")
        df.to_excel(writer, index=False, sheet_name=main_sheet_name)

        # 2. D열 고유값별 시트 생성
        print(f"\n  D열 고유값별 시트 생성 중...")
        for d_value in unique_d_values:
            # 필터링된 데이터
            filtered_df = df[df[d_column_name] == d_value].copy()

            # 시트명 (Excel 시트명 제한: 31자, 특수문자 제거)
            sheet_name = str(d_value)[:31]
            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('*', '_')
            sheet_name = sheet_name.replace('[', '_').replace(']', '_').replace(':', '_')
            sheet_name = sheet_name.replace('?', '_')

            # 시트 생성
            filtered_df.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"    시트 '{sheet_name}': {len(filtered_df)}건")

        # 워크북 가져오기
        wb = writer.book

        # 모든 시트에 동일한 서식 적용
        print(f"\n  모든 시트에 서식 적용 중...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 열 너비 적용
            column_widths = {
                'A': 6.66, 'B': 9.83, 'C': 8.83, 'D': 13.83, 'E': 12.16,
                'F': 21.33, 'G': 14.5, 'H': 13.0, 'I': 13.0, 'J': 17.5,
                'K': 8.5, 'L': 13.0, 'M': 13.0, 'N': 21.83, 'O': 28.33,
                'P': 17.16, 'Q': 10.0, 'R': 10.16, 'S': 10.0, 'T': 11.5
            }

            for col_letter, width in column_widths.items():
                if col_letter_to_index(col_letter) <= len(df.columns):
                    ws.column_dimensions[col_letter].width = width

            # 전체 셀 가운데 정렬 + 텍스트 자동 줄바꿈
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center_alignment

            # 제목행 스타일 (1행)
            bold_font = Font(bold=True)
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = gray_fill

            # 우선순위별 행 배경색 적용 (2행부터)
            priority_colors = {
                '1순위': 'FFFF00',  # 노란색
                '2순위': 'F7B9AF'   # 분홍색
            }

            priority_col_idx = len(df.columns)  # 마지막 열 (점검순위)

            for row_idx in range(2, ws.max_row + 1):
                priority = ws.cell(row=row_idx, column=priority_col_idx).value

                if priority in priority_colors:
                    color = priority_colors[priority]
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = fill

            # 모든 셀에 테두리 적용
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # 틀 고정 (1행 고정)
            ws.freeze_panes = 'A2'

            # 자동 필터 적용
            ws.auto_filter.ref = ws.dimensions

    print(f"\n  저장 완료: {output_filename}")
    print(f"  총 시트 수: {len(unique_d_values) + 1}개 (메인 1개 + D열 고유값 {len(unique_d_values)}개)")
    print("=" * 60 + "\n")

    return output_filepath
